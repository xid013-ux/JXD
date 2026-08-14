# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
OUT='/home/user/JXD/output/金源/无锡金源_公司画像底稿.xlsx'
C=Alignment(horizontal='center',vertical='center',wrap_text=True)
L=Alignment(horizontal='left',vertical='top',wrap_text=True)
HDR=PatternFill('solid',fgColor='123F63'); HI=PatternFill('solid',fgColor='D6E9D6')
HF=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF'); BF=Font(name='微软雅黑',size=10)
thin=Side(style='thin',color='BBBBBB'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook(); wb.remove(wb.active)
def mk(name,title,headers,rows,widths,hi=None):
    ws=wb.create_sheet(name); ws['A1']=title
    ws['A1'].font=Font(name='微软雅黑',size=12,bold=True,color='123F63')
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(headers))
    ws['A1'].alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[1].height=22
    for i,h in enumerate(headers,1):
        c=ws.cell(row=2,column=i,value=h); c.fill=HDR; c.font=HF; c.alignment=C; c.border=BD
    for r,row in enumerate(rows,3):
        for i,v in enumerate(row,1):
            c=ws.cell(row=r,column=i,value=v); c.font=BF; c.alignment=C; c.border=BD
        if hi and str(row[hi-1]).startswith('★'):
            for i in range(1,len(headers)+1): ws.cell(row=r,column=i).fill=HI
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A3'

mk('01_工商与资质','一、工商登记与资质（资料来源：企查查公示信息，2026年8月6日更新）',
 ['项目','内容'],
 [['企业名称','无锡金源半导体科技有限公司'],
  ['统一社会信用代码','91320211MA1Y2PWY0J'],
  ['法定代表人','刘亚（关联企业10家）'],
  ['成立日期','2019年3月15日'],
  ['登记状态','在业'],
  ['注册资本','2,780.932万元'],
  ['实缴资本','2,730.932万元'],
  ['企业类型','有限责任公司'],
  ['注册地址','无锡市滨湖区胡埭镇孟村路1号'],
  ['登记机关','无锡市滨湖区数据局'],
  ['国标行业','研究和试验发展（M73）'],
  ['参保人数','121人（2025年报）'],
  ['资质标签','高新技术企业、科技型中小企业、专精特新中小企业、企业技术中心'],
  ['官网','www.jinyuansemi.com'],
  ['经营范围（摘要）','半导体器件专用设备制造与销售、密封件制造、特种陶瓷制品制造与销售、'
   '电子专用材料研发制造销售、金属制品研发与制造、阀门和旋塞研发制造销售、'
   '技术玻璃制品制造、技术进出口与货物进出口'],
 ],[18,76])

mk('02_产品线','二、产品体系（资料来源：公司官网产品中心，2026年8月）',
 ['标记','产品线','官网型号或类别','观察到的信息','与可比公司的对应'],
 [
 ['','O型圈（密封圈）','自有品牌CeaLution®，型号包括F7068、F7072、F5076T、F7070、F5075、F9075',
  '同一品牌下按型号区分，产品外观颜色不同（棕色、棕黄色、黑色、白色），'
  '指向不同配方或不同耐受工况',
  '三家主要可比公司均不覆盖该品类；'
  '境内上市公司中唯万密封涉及半导体用全氟醚橡胶密封件'],
 ['','密封件门板','方形门板、圆形门板',
  '门板为腔体开合部位的密封结构件，图片显示为金属基体配密封圈的组合件',
  '属结构件与密封件的组合，托伦斯的腔体类产品与之相邻'],
 ['★','精密气体喷头','LAM AHM 300mm、LAM STRATA 300mm、LAM Q-STRATA 300mm',
  '产品直接以设备厂商机型命名，均为300mm晶圆制程；'
  '型号对应泛林集团（Lam Research）的机台系列',
  '对应托伦斯的匀气盘、气体分布盘；'
  '托伦斯问询回复披露匀气盘2023年毛利率27.44%、气体分布盘20.05%'],
 ['','加热盘','官网该栏目显示"暂无信息"',
  '产品线已在导航中列出，但未上架具体型号',
  '对应托伦斯的加热器（2023年毛利率14.39%，2022年为-18.68%）、'
  '珂玛科技与臻宝科技的陶瓷加热器'],
 ],[6,16,32,44,44],hi=1)

mk('03_待核实','三、需在拜访中当面了解的事项（公开渠道无法取得）',
 ['序号','事项','为什么要问'],
 [
 ['1','四条产品线各自的收入占比与毛利率',
  '可比公司的毛利率按材料属性分层明显：非金属类约50%、金属类约25%、'
  '含模组类约16%。金源四条线材料属性不同，需了解各自水平'],
 ['2','精密气体喷头以LAM机型命名，是面向晶圆厂的替换备件，'
  '还是向设备厂配套供货',
  '两种模式的客户结构、订单节奏与认证周期完全不同。'
  '臻宝科技的定位为直接供应晶圆厂，托伦斯的客户为北方华创、中微公司等设备厂'],
 ['3','是否已进入国内主要晶圆厂或设备厂的合格供应商名录，验证周期多长',
  '客户导入情况是审核中的核心关注点之一'],
 ['4','加热盘产品线的进展与规划',
  '官网未上架具体型号，需了解处于何种阶段'],
 ['5','O型圈的原材料来源，全氟醚橡胶是自制混炼还是外购',
  '臻宝科技第二轮问询即被追问原材料通过委托加工与成品采购取得、'
  '供应商规模较小、缺乏可比采购价格的问题'],
 ['6','近三年营业收入、净利润与产能利用率',
  '非上市公司，公开渠道无数据'],
 ['7','股权结构、历史融资与是否存在对赌安排',
  '托伦斯首轮问询设有历史沿革、对赌协议两个专项'],
 ['8','研发投入水平与核心技术人员构成',
  '可比公司研发费用率2025年平均5.71%，区间为1.56%至9.32%'],
 ],[6,40,64])
wb.save(OUT); print('saved',OUT)
