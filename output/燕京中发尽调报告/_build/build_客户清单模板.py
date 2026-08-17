# -*- coding: utf-8 -*-
"""2-4-2 / 1-8-5 客户清单填报模板（简洁版）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F = 'Arial'
TIT  = Font(name=F, size=13, bold=True)
H    = Font(name=F, size=10, bold=True)
BOLD = Font(name=F, size=10, bold=True)
N    = Font(name=F, size=10)
NOTE = Font(name=F, size=9, color='808080')
EX   = Font(name=F, size=10, color='808080', italic=True)
BLUE = Font(name=F, size=10, color='0000FF')
GREEN= Font(name=F, size=10, color='008000')

GREY = PatternFill('solid', fgColor='F2F2F2')          # 唯一底色：表头
thin = Side(style='thin', color='D9D9D9')
med  = Side(style='thin', color='808080')
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)
HBOX = Border(left=thin, right=thin, top=med, bottom=med)
TOPL = Border(left=thin, right=thin, top=med, bottom=thin)
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)

PERIODS = ['2024年度', '2025年度', '2026年1-6月']
REV_ZF = {'2024年度': 212656418.65, '2025年度': 434358616.85, '2026年1-6月': 208146233.34}
REV_XT = {'2024年度': 5163572.58,   '2025年度': 100430167.78, '2026年1-6月': 29375817.27}

COLS = [
    ('期间', 12), ('序号', 5), ('客户名称（全称）', 32), ('统一社会信用代码', 21),
    ('客户类型', 12), ('销售模式', 13), ('经销区域', 13), ('主要销售产品', 20),
    ('合作起始时间', 12), ('本期销售金额\n（不含税，元）', 16), ('占营业收入\n比例', 12),
    ('本期回款金额\n（含税，元）', 16), ('期末应收账款\n余额（元）', 15),
    ('期末合同负债\n余额（元）', 15), ('是否存在\n第三方回款', 11),
    ('第三方回款方名称及金额', 24), ('是否关联方', 11),
    ('关联关系说明\n（是否为股东、董监高、员工及其近亲属持股或任职）', 32), ('备注', 16),
]
NC = len(COLS)
NROW = 20
MONEY = (10, 12, 13, 14)
CTRCOL = (1, 2, 5, 6, 9, 11, 15, 17)

def build_sheet(wb, title, entity, revmap, example):
    ws = wb.create_sheet(title)
    ws['A1'] = f'{entity}　主要客户清单'
    ws['A1'].font = TIT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)

    ws['A2'] = ('各期至少填列前20名客户，且合计占当期营业收入不低于50%；未达标时请在该期区块下方插入行继续补充。'
                '“占营业收入比例”列及各“前20名合计”行为公式，请勿修改；其余各栏请填写。')
    ws['A2'].font = NOTE
    ws['A2'].alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
    ws.row_dimensions[2].height = 26

    # 营业收入（占比公式的唯一数据源）
    ws['A3'] = '当期营业收入（元）'
    ws['A3'].font = BOLD
    REVCELL, col = {}, 2
    for p in PERIODS:
        ws.cell(3, col, p).font = N
        ws.cell(3, col).alignment = Alignment(horizontal='right', vertical='center')
        v = ws.cell(3, col + 1, revmap[p]); v.font = BLUE; v.number_format = '#,##0.00'
        REVCELL[p] = f'${get_column_letter(col + 1)}$3'
        col += 2
    ws.cell(3, col, '蓝色数字取自财务报表，如与账面不符请直接修改').font = NOTE

    hr = 5
    for i, (name, w) in enumerate(COLS, start=1):
        c = ws.cell(hr, i, name)
        c.font, c.fill, c.alignment, c.border = H, GREY, CTR, HBOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hr].height = 40
    ws.freeze_panes = ws.cell(hr + 1, 3)

    er = hr + 1
    for i, v in enumerate(example, start=1):
        c = ws.cell(er, i, v)
        c.font, c.border = EX, BOX
        c.alignment = CTR if i in CTRCOL else WRAP
        if i in MONEY: c.number_format = '#,##0.00'
    ws.cell(er, 19, '← 示例行，填报时请删除')

    r, blocks = er + 1, []
    for p in PERIODS:
        ws.cell(r, 1, p).font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        r += 1
        first = r
        for k in range(1, NROW + 1):
            ws.cell(r, 1, p).font = N
            ws.cell(r, 2, k).font = N
            for c in range(1, NC + 1):
                cell = ws.cell(r, c)
                cell.border, cell.font = BOX, N
                cell.alignment = CTR if c in CTRCOL else WRAP
                if c == 11:
                    cell.value = f'=IF(J{r}="","",J{r}/{REVCELL[p]})'
                    cell.number_format = '0.00%'
                elif c in MONEY:
                    cell.number_format = '#,##0.00'
            r += 1
        last = r - 1
        ws.cell(r, 3, '前20名合计').font = BOLD
        for c in MONEY:
            L = get_column_letter(c)
            cc = ws.cell(r, c, f'=SUM({L}{first}:{L}{last})')
            cc.number_format, cc.font = '#,##0.00', BOLD
        cc = ws.cell(r, 11, f'=IF(J{r}=0,"",J{r}/{REVCELL[p]})')
        cc.number_format, cc.font = '0.00%', BOLD
        ws.cell(r, 15, '是否达标').font = BOLD
        cc = ws.cell(r, 16, f'=IF(J{r}=0,"未填",IF(K{r}>=0.5,"达标","未达标，请补充客户"))')
        cc.font = BOLD
        for c in range(1, NC + 1):
            ws.cell(r, c).border = TOPL
            if ws.cell(r, c).alignment is None or c in CTRCOL:
                ws.cell(r, c).alignment = CTR
        blocks.append((p, r))
        r += 2

    # 勾稽核对
    ws.cell(r, 1, '勾稽核对').font = BOLD
    r += 1
    for i, x in enumerate(['期间', '当期营业收入（元）', '前20名合计（元）', '合计占比',
                           '前20名以外客户合计（元）'], start=1):
        c = ws.cell(r, i, x); c.font, c.fill, c.alignment, c.border = H, GREY, CTR, HBOX
    r += 1
    for p, sub in blocks:
        ws.cell(r, 1, p).font = N; ws.cell(r, 1).alignment = CTR
        for i, (val, fmt, fnt) in enumerate([
                (f'={REVCELL[p]}', '#,##0.00', GREEN),
                (f'=J{sub}', '#,##0.00', GREEN),
                (f'=IF(C{r}=0,"",C{r}/B{r})', '0.00%', N),
                (f'=IF(C{r}=0,"",B{r}-C{r})', '#,##0.00', N)], start=2):
            c = ws.cell(r, i, val); c.number_format, c.font = fmt, fnt
        for i in range(1, 6):
            ws.cell(r, i).border = BOX
        r += 1
    ws.cell(r, 1, '注：绿色字体为引用本表其他单元格的公式。').font = NOTE

    n_end = er + 1 + (NROW + 2) * 3
    for f1, rng in [('"经销商,直销客户,电商平台,商超,其他"', f'E{er+1}:E{n_end}'),
                    ('"买断经销,委托代销,直销,平台入仓,平台代销,其他"', f'F{er+1}:F{n_end}'),
                    ('"是,否"', f'O{er+1}:O{n_end}'), ('"是,否"', f'Q{er+1}:Q{n_end}')]:
        dv = DataValidation(type='list', formula1=f1, allow_blank=True)
        ws.add_data_validation(dv); dv.add(rng)


wb = openpyxl.Workbook()
ws = wb.active; ws.title = '填报说明'
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 104
ws['A1'] = '燕京中发IPO项目　客户清单填报说明'
ws['A1'].font = Font(name=F, size=14, bold=True)
ws.merge_cells('A1:B1')

INFO = [
 ('清单编号', '2-4-2（北京燕京中发生物技术有限公司）、1-8-5（燕京中发生物技术（邢台）有限公司）'),
 ('填报范围', '两家主体分别填报，各填三个期间：2024年度、2025年度、2026年1-6月。'),
 ('填报口径', '1、销售金额为不含税口径，与当期利润表营业收入可勾稽；回款金额为含税口径，指本期实际收到的款项。\n'
             '2、客户名称填工商登记全称，与销售合同、发票名称一致；同一客户各期须用同一名称。\n'
             '3、同一集团下不同法人主体分别列示，并在备注注明所属集团。\n'
             '4、各期按销售金额由大到小排列。'),
 ('比例要求', '各期前列客户销售金额合计占当期营业收入不得低于50%。前20名合计未达50%的，请继续补充客户，'
             '直至“是否达标”显示“达标”。'),
 ('哪些要填', '除“占营业收入比例”列、各“前20名合计”行及“勾稽核对”区为自动计算外，其余各栏均需填写。'
             '各表第3行的营业收入为蓝色字体的预填数据，请公司复核，如与账面不符可直接修改。'),
 ('重点一', '第三方回款：如存在合同签订方、发票开具方与实际汇款方不一致的情形（例如由经销商的经营者、'
          '股东个人账户付款），务必如实填列付款方名称及金额。该事项为IPO审核重点核查事项。'),
 ('重点二', '关联关系：请核查客户及其股东、实际控制人、主要经办人员，与公司的股东、董事、监事、高级管理人员、'
          '员工及其近亲属之间是否存在持股、任职、亲属或其他关联关系，并如实填列。'),
 ('重点三', '两家主体的客户名称口径必须统一，项目组将据以分析客户重合情况。'),
 ('数据来源', '燕京中发营业收入：2024年212,656,418.65元、2025年434,358,616.85元、2026年1-6月208,146,233.34元，'
            '取自2024年度审计报告、2025年度财务报表及2026年6月30日利润表。\n'
            '中发邢台营业收入：2024年5,163,572.58元、2025年100,430,167.78元、2026年1-6月29,375,817.27元，'
            '取自其2024、2025年度审计报告及2026年6月30日利润表。'),
 ('反馈', '请连同“2-4-3 上述客户的销售合同”一并反馈。某栏确无法提供的，请填“无”或“不适用”并说明原因，不要留空。'),
]
r = 3
for k, v in INFO:
    a = ws.cell(r, 1, k); a.font = BOLD; a.alignment = Alignment(vertical='top'); a.border = BOX
    b = ws.cell(r, 2, v); b.font = N; b.alignment = WRAP; b.border = BOX
    ws.row_dimensions[r].height = max(26, 13.5 * (v.count('\n') + 1) + 13)
    r += 1

EX_ZF = ['2025年度', 1, '示例：××省××商贸有限公司', '91XXXXXXXXXXXXXXXX', '经销商', '买断经销', '华北区',
         '肠溶型纳豆胶囊、燕京纳福多肽', '2021-03', 12345678.90, None, 13950000.00, 0.00, 1200000.00,
         '是', '××（该经销商股东）付款 300,000.00 元', '否', '无', '']
EX_XT = ['2025年度', 1, '示例：××省××生物科技有限公司', '91XXXXXXXXXXXXXXXX', '经销商', '买断经销', '华东区',
         '纳福多肽、有机全脂羊乳粉', '2024-08', 8765432.10, None, 9905000.00, 0.00, 500000.00,
         '否', '无', '否', '无', '']

build_sheet(wb, '客户清单-燕京中发', '北京燕京中发生物技术有限公司', REV_ZF, EX_ZF)
build_sheet(wb, '客户清单-中发邢台', '燕京中发生物技术（邢台）有限公司', REV_XT, EX_XT)

ws = wb.create_sheet('客户重合度分析')
ws['A1'] = '发行人与中发邢台客户重合情况分析'
ws['A1'].font = TIT; ws.merge_cells('A1:J1')
ws['A2'] = '本表由项目组根据前两表填报结果编制，用于同业竞争事项论证，公司无需填写。'
ws['A2'].font = NOTE; ws.merge_cells('A2:J2')
hh = ['序号', '客户名称（全称）', '统一社会信用代码', '燕京中发\n2024年度', '燕京中发\n2025年度',
      '燕京中发\n2026年1-6月', '中发邢台\n2024年度', '中发邢台\n2025年度', '中发邢台\n2026年1-6月', '说明']
for i, (x, w) in enumerate(zip(hh, [5, 32, 21, 15, 15, 15, 15, 15, 15, 28]), start=1):
    c = ws.cell(4, i, x); c.font, c.fill, c.alignment, c.border = H, GREY, CTR, HBOX
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[4].height = 34
for r in range(5, 35):
    for i in range(1, 11):
        c = ws.cell(r, i); c.border, c.font = BOX, N
        c.alignment = CTR if i == 1 else WRAP
        if 4 <= i <= 9: c.number_format = '#,##0.00'
ws.freeze_panes = 'C5'
ws.cell(35, 3, '合计').font = BOLD
for i in range(4, 10):
    L = get_column_letter(i)
    c = ws.cell(35, i, f'=SUM({L}5:{L}34)')
    c.number_format, c.font, c.border = '#,##0.00', BOLD, TOPL
ws.cell(35, 3).border = TOPL

wb.save('2-4-2 客户清单填报模板.xlsx')
print('已生成（简洁版）')
