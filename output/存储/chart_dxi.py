# -*- coding: utf-8 -*-
import matplotlib; matplotlib.use('Agg')
matplotlib.rcParams['font.sans-serif']=['WenQuanYi Zen Hei']
matplotlib.rcParams['axes.unicode_minus']=False
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
from datetime import datetime

P=['#123F63','#2E6C97','#5B93B8','#8FB6CE','#C3D6E3','#E3EBF1']
wb=load_workbook('/root/.claude/uploads/b33c1a84-553f-5c48-a1f8-b4ca1cd8f155/d73b6884-DXI__.xlsx',data_only=True)
ws=wb['DXI指数']
data=[]
for r in ws.iter_rows(min_row=8, values_only=True):
    d,v=r[0],r[1]
    if isinstance(d,datetime) and isinstance(v,(int,float)): data.append((d,v))
data.sort()
xs=[d for d,_ in data]; ys=[v for _,v in data]

fig,ax=plt.subplots(figsize=(10,4.6))
ax.plot(xs,ys,color=P[0],lw=1.6)
ax.fill_between(xs,ys,color=P[3],alpha=.28)
ax.set_yscale('log')
ax.set_ylabel('DXI指数（点，对数刻度）')
for s in ('top','right'): ax.spines[s].set_visible(False)
ax.grid(axis='y',linestyle='--',color='#DDDDDD'); ax.set_axisbelow(True)
ax.yaxis.set_major_formatter(FuncFormatter(lambda v,p: f'{v:,.0f}'))

ann=[('2020-03-19','疫情初期低位',18,14),('2021-07-15','2021年周期高点',0,16),
     ('2023-12-29','下行周期底部',0,-26),('2024-12-31','AI需求驱动回升',-52,14),
     ('2026-08-11','2026-08-11\n965,148点',-72,-6)]
for ds,lab,dx,dy in ann:
    m=[(d,v) for d,v in data if d.strftime('%Y-%m-%d')==ds]
    if not m: continue
    d,v=m[0]
    ax.scatter([d],[v],s=26,color=P[0],zorder=5)
    ax.annotate(f'{lab}\n{v:,.0f}' if '\n' not in lab else lab,(d,v),
                textcoords='offset points',xytext=(dx,dy),fontsize=8.5,color='#333333')
fig.tight_layout()
fig.savefig('/home/user/JXD/output/存储/charts/fig_dxi.png',dpi=200)
print('图已生成')

# 年末值 + 涨跌
import collections
by=collections.OrderedDict()
for d,v in data: by[d.year]=(d,v)
print('\n年末值：')
prev=None
for y,(d,v) in by.items():
    print(f'  {d.date()}  {v:>12,.2f}' + (f'   {v/prev-1:+.1%}' if prev else ''))
    prev=v
