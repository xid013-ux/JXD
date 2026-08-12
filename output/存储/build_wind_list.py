# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/home/user/JXD/output/存储/半导体存储行业研究_取数清单.xlsx'
C = Alignment(horizontal='center', vertical='center', wrap_text=True)
L = Alignment(horizontal='left', vertical='center', wrap_text=True)
HDR = PatternFill('solid', fgColor='123F63')
P0 = PatternFill('solid', fgColor='F5C6C6')
P1 = PatternFill('solid', fgColor='FCE4C4')
P2 = PatternFill('solid', fgColor='DCE6F1')
HF = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BF = Font(name='微软雅黑', size=10)
thin = Side(style='thin', color='BBBBBB')
BD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); wb.remove(wb.active)

def mk(name, title, headers, rows, widths, prio_col=1):
    ws = wb.create_sheet(name)
    ws['A1'] = title
    ws['A1'].font = Font(name='微软雅黑', size=13, bold=True, color='123F63')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].alignment = L; ws.row_dimensions[1].height = 24
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HDR; c.font = HF; c.alignment = C; c.border = BD
    for r, row in enumerate(rows, 3):
        for i, v in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BF; c.alignment = C; c.border = BD
        p = row[prio_col - 1]
        fill = P0 if p == 'P0' else (P1 if p == 'P1' else P2)
        ws.cell(row=r, column=prio_col).fill = fill
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'
    return ws

# ---------- 说明 ----------
ws = wb.create_sheet('00_怎么用这张表')
lines = [
    '半导体存储行业研究 —— 取数清单',
    '',
    '编制时点：2026年8月12日',
    '',
    '为什么需要这张表：',
    'AI侧能做的只有公开网页检索，拿到的基本都是媒体与行业站点的转述数据。',
    'Wind、巨潮资讯、交易所网站在当前网络环境下均无法访问，原始文件无法下载。',
    '按本部门规范，转述出处的数据不能直接写入正文，必须回到原始发布方核实。',
    '因此第四步回源需要由人从Wind与公开披露渠道取得原始材料。',
    '',
    '优先级说明：',
    'P0  没有这些材料，第二章（封测环节）与第八章（深科技）无法动笔',
    'P1  影响第三章（市场规模）与第四章（竞争格局）的数据质量',
    'P2  影响第五、六、七章（代表企业、资本化、国泰海通关联），可后补',
    '',
    '交付形式：',
    'PDF原始文件直接发我即可；Wind导出的数据存成Excel发我，',
    '每张表请保留Wind的指标名称与时间标签，不要手工改列名，便于回源留痕。',
    '',
    '关于拿不到的情况：',
    '任何一项如果Wind与公开渠道都查不到，直接告诉我"查不到"即可。',
    '按规范这类内容会整条从正文删除，不会写成"据媒体报道"或标注待核实。',
]
for i, t in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = Font(name='微软雅黑', size=13, bold=True, color='123F63') if i == 1 \
        else Font(name='微软雅黑', size=10)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 110

# ---------- A. 原始文件 ----------
mk('01_需要下载的原始文件', 'A、需要下载的原始文件（PDF直接发我）',
   ['优先级', '文件', '获取渠道', '要从里面取什么', '对应章节'],
   [
    ['P0', '深科技（000021.SZ）2025年年度报告',
     '巨潮资讯网 或 Wind F9→公告',
     '分行业与分产品的主营业务收入构成（收入、成本、毛利率）；'
     '主要子公司情况中深圳沛顿、合肥沛顿存储的营收与净利润；'
     '产能与在建工程；主要客户集中度', '第二章、第五章、第八章'],
    ['P0', '深科技 2026年第一季度报告', '巨潮资讯网 或 Wind',
     '最新营收与利润；经营情况说明中关于存储封测的表述', '第八章'],
    ['P0', '深科技 关于沛顿科技扩产的对外投资公告',
     '巨潮资讯网（2026年，约14.7亿元项目）',
     '投资额、产能规模、建设周期、资金来源、项目主体', '第二章、第八章'],
    ['P0', '深科技 2024年、2023年年度报告',
     '巨潮资讯网', '主营构成的历史数据，用于做三年趋势', '第八章'],
    ['P1', '长鑫科技 招股说明书（注册稿）',
     '上交所科创板项目动态',
     '行业章节的市场规模与份额数据（含引用的机构与口径）；'
     '产能、技术节点、客户结构；同行业可比公司；募投项目', '第三章、第四章、第五章'],
    ['P1', '长鑫科技 发行结果公告、上市公告书',
     '上交所指定披露渠道', '发行价、募资额、战略配售、保荐机构', '第六章'],
    ['P1', '长鑫科技 审核问询函回复（两轮）',
     '上交所科创板项目动态',
     '监管关注的行业口径问题；公司对份额与产能的解释', '第三章、第四章'],
    ['P2', '兆易创新、江波龙、佰维存储、德明利 2025年年度报告',
     '巨潮资讯网', '各自的产品结构、收入、毛利率、客户', '第五章'],
    ['P2', '长电科技、通富微电、华天科技 2025年年度报告',
     '巨潮资讯网', '封测业务收入构成；存储封测相关表述与产能', '第二章、第四章'],
   ],
   [8, 30, 26, 46, 20])

# ---------- B. Wind ----------
mk('02_Wind需要拉的数据', 'B、Wind 需要拉的数据（导出Excel发我）',
   ['优先级', '要什么数据', 'Wind里大致的位置', '范围与口径要求', '对应章节'],
   [
    ['P0', '深科技主营业务构成',
     'F9 → 公司资料 → 主营构成（或 数据浏览器）',
     '2022—2025年报口径，分行业与分产品，'
     '含营业收入、营业成本、毛利率三项；如有2026半年报一并取',
     '第八章、第二章'],
    ['P0', '深科技财务指标时间序列',
     'F9 → 财务分析 / 数据浏览器',
     '2020—2025年度 + 2026Q1：营业收入、归母净利润、毛利率、净利率、'
     '研发费用、资产负债率、经营性现金流', '第八章'],
    ['P1', '全球存储市场规模与份额',
     '行业数据库；或研报数据库中引用TrendForce／集邦／Omdia的原始图表',
     '2020—2025年DRAM与NAND分别的市场规模与厂商份额，'
     '务必记录数据发布机构、统计对象（销售额还是位元）与时点',
     '第三章、第四章'],
    ['P1', '存储器价格指数',
     'EDB宏观经济数据库（如有DXI指数或DRAM/NAND合约价、现货价）',
     '2020年至今月度或周度，标明品类与规格；'
     '用于画价格周期图', '第三章'],
    ['P1', '半导体封装测试行业规模',
     '行业数据库',
     '需区分：全部芯片封测总盘 与 存储封测细分；'
     '若Wind只有总盘，请注明，不要用总盘代替细分', '第二章'],
    ['P2', '存储产业链上市公司清单与财务',
     '板块 → 概念板块（存储芯片／存储器）→ 导出成分',
     '公司名称、代码、上市板、主营业务、2025年营收与归母净利润、'
     '所处产业链环节；A股、H股、A+H都要覆盖', '第五章、第六章'],
    ['P2', '存储产业链公司IPO与再融资信息',
     '股票发行数据集',
     '上市日期、发行价、募资额、保荐机构（区分独家与联席）、'
     '项目类型（IPO／再融资／可转债）；含合并前国泰君安、海通证券的项目',
     '第六章、第七章'],
    ['P2', '在审、已受理、已递表的存储公司',
     'IPO进程／发行监管数据集',
     '公司名称、板块、受理日期、当前状态、保荐机构',
     '第六章'],
    ['P2', '辅导备案中的存储公司',
     '辅导备案数据集',
     '公司名称、辅导备案日期、辅导券商、所在地证监局', '第六章'],
    ['P2', '国泰海通相关的存储产业链项目',
     '按承销保荐机构筛选；股权投资可查股东穿透',
     '国泰海通、国泰君安、海通证券作为保荐机构的存储产业链项目；'
     '国泰君安证裕、海通开元、国泰君安创新投资等主体的股权投资',
     '第七章'],
   ],
   [8, 26, 30, 46, 18])

# ---------- C. 我这边自己做 ----------
mk('03_我这边自己做的', 'C、不用麻烦你的部分（我用公开检索完成，但结论仍需上述原始文件印证）',
   ['优先级', '事项', '我的做法', '局限'],
   [
    ['P1', '产业链结构梳理与技术路线',
     '公开技术资料与行业综述整理',
     '不涉及具体数值，风险较低'],
    ['P1', '政策梳理', '主管部门与地方政府公开文件检索',
     '政策原文一般可检索到，但正式引用仍以文件全称与发文字号为准'],
    ['P2', '海外原厂对照（三星、SK海力士、美光）',
     '各公司公开财报数据检索',
     '财报原文下载受限，数值需以你能拉到的Wind数据交叉印证'],
    ['P2', '报告的框架、写作与排版', '按部门规范生成Word与Excel',
     '本环境无法渲染PDF，最终版式需你在Word中确认'],
   ],
   [8, 26, 40, 40])

wb.save(OUT)
print('saved', OUT)
